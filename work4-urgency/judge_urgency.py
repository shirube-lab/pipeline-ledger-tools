"""Judge sewer-span urgency (緊急度Ⅰ/Ⅱ/Ⅲ) from TV-camera survey results.

Implements the reference judgment method shown in MLIT's
"下水道長寿命化支援制度に関する手引き(案)" (H21, 2009) tables 2.3-2.9,
cross-checked against the H25 successor documents. The thresholds and
rules live in a criteria profile (criteria/default_h21.json), not in
code, so municipality-specific profiles can be added without touching
the engine.

Input (two CSVs, modeled on real survey record sheets):
    span_master.csv    span_id, mh_up, mh_down, material, diameter_mm,
                       length_m, n_pipes
    defect_records.csv span_id, pipe_no, distance_m, item, rank, note
                       - span-level items (腐食/たるみ): rank A/B/C, pipe_no empty
                       - pipe-level items (破損 etc.): rank a/b/c

Output:
    output/span_urgency.csv    per-span urgency with reasoning
    output/urgency_report.xlsx summary / per-span / defects / criteria

Usage:
    python judge_urgency.py [--criteria criteria/default_h21.json]
                            [--span span_master.csv] [--defect defect_records.csv]

Note: this tool judges POST-survey urgency (診断). It is not a
pre-survey screening tool — see README for the terminology distinction.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

HERE = Path(__file__).parent

# Severity orders: index 0 is worst
PIPE_SEVERITY = ["a", "b", "c"]
SPAN_SEVERITY = ["A", "B", "C"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)

URGENCY_COLORS = {"Ⅰ": "C00000", "Ⅱ": "ED7D31", "Ⅲ": "FFC000"}


def load_criteria(path: Path) -> dict:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def validate(span_df: pd.DataFrame, defect_df: pd.DataFrame, criteria: dict) -> pd.DataFrame:
    """Validate defect records against the criteria profile.

    Returns a DataFrame of validation errors (empty when clean).
    Invalid records are excluded from judgment by the caller.
    """
    span_items = set(criteria["span_items"])
    pipe_items = set(criteria["pipe_items"])
    materials = dict(zip(span_df["span_id"], span_df["material"]))
    errors = []

    for idx, r in defect_df.iterrows():
        sid, item, rank = r["span_id"], r["item"], str(r["rank"])
        if sid not in materials:
            errors.append({"row": idx, "span_id": sid, "issue": "span_master に存在しないスパンID"})
            continue
        if item in span_items:
            if rank not in SPAN_SEVERITY:
                errors.append({"row": idx, "span_id": sid,
                               "issue": f"スパン評価項目「{item}」のランクは A/B/C({rank} は不正)"})
        elif item in pipe_items:
            if rank not in PIPE_SEVERITY:
                errors.append({"row": idx, "span_id": sid,
                               "issue": f"管1本評価項目「{item}」のランクは a/b/c({rank} は不正)"})
            else:
                mat = materials[sid]
                defined = criteria["pipe_items"][item]["ranks"].get(mat, {})
                if defined.get(rank) is None:
                    errors.append({"row": idx, "span_id": sid,
                                   "issue": f"「{item}」ランク {rank} は管種 {mat} では定義されない(H21 表2.3)"})
        else:
            errors.append({"row": idx, "span_id": sid, "issue": f"未定義の調査項目「{item}」"})

    return pd.DataFrame(errors, columns=["row", "span_id", "issue"])


def worst(ranks: list[str], order: list[str]) -> str | None:
    present = [r for r in ranks if r in order]
    if not present:
        return None
    return order[min(order.index(r) for r in present)]


def dedupe_same_location(pipe_defects: pd.DataFrame) -> pd.DataFrame:
    """Keep only the worst-ranked defect per location (H21 表2.7 備考③).

    A "location" is the same pipe at the same distance. The kept row
    gets a flag so the report can show what was collapsed.
    """
    if pipe_defects.empty:
        return pipe_defects.assign(dedup_note="")
    df = pipe_defects.copy()
    df["_sev"] = df["rank"].map({r: i for i, r in enumerate(PIPE_SEVERITY)})
    df = df.sort_values("_sev")
    keys = ["span_id", "pipe_no", "distance_m"]
    dup_sizes = df.groupby(keys)["rank"].transform("size")
    kept = df.drop_duplicates(subset=keys, keep="first").copy()
    kept["dedup_note"] = ""
    collapsed = dup_sizes.loc[kept.index] > 1
    kept.loc[collapsed, "dedup_note"] = "同一箇所の複数不良 → 最上位のみ計上(備考③)"
    return kept.drop(columns="_sev")


def rate_rank(n_pipes: int, per_pipe_worst: list[str], thresholds: dict) -> tuple[str | None, dict]:
    """Span rank from defect occurrence rates (H21 表2.6-2.7).

    per_pipe_worst: each surveyed pipe's worst pipe-item rank. A pipe is
    counted once, at its worst rank ("合計本数" counts pipes, and 備考③
    already collapses multiple defects at one location).
    """
    n_a = sum(1 for r in per_pipe_worst if r == "a")
    n_b = sum(1 for r in per_pipe_worst if r == "b")
    n_c = sum(1 for r in per_pipe_worst if r == "c")
    a_pct = n_a / n_pipes * 100
    ab_pct = (n_a + n_b) / n_pipes * 100
    abc_pct = (n_a + n_b + n_c) / n_pipes * 100
    rates = {"n_a": n_a, "n_b": n_b, "n_c": n_c,
             "a_pct": round(a_pct, 1), "ab_pct": round(ab_pct, 1), "abc_pct": round(abc_pct, 1)}

    th = thresholds
    if n_a + n_b + n_c == 0:
        return None, rates
    if a_pct >= th["A"]["a_pct_min"] or ab_pct >= th["A"]["ab_pct_min"]:
        return "A", rates
    if n_a > 0 or n_b > 0:
        return "B", rates
    if abc_pct >= th["B"]["abc_pct_min"]:
        return "B", rates
    return "C", rates


def judge_span(span: pd.Series, defects: pd.DataFrame, criteria: dict) -> dict:
    """Judge one span. Returns a result dict including the reasoning text."""
    span_items = set(criteria["span_items"])
    pipe_item_defs = criteria["pipe_items"]
    immediate_items = {k for k, v in pipe_item_defs.items() if v.get("immediate_A_on_a")}

    span_level = defects[defects["item"].isin(span_items)]
    pipe_level = defects[defects["item"].isin(pipe_item_defs)]

    corrosion = worst(list(span_level[span_level["item"] == "腐食"]["rank"]), SPAN_SEVERITY)
    sag = worst(list(span_level[span_level["item"] == "たるみ"]["rank"]), SPAN_SEVERITY)

    deduped = dedupe_same_location(pipe_level)
    per_pipe = (
        deduped.groupby("pipe_no")["rank"]
        .apply(lambda s: worst(list(s), PIPE_SEVERITY))
        .dropna()
        .tolist()
    )
    r_rank, rates = rate_rank(int(span["n_pipes"]), per_pipe, criteria["rate_thresholds"])

    immediate_hits = deduped[(deduped["item"].isin(immediate_items)) & (deduped["rank"] == "a")]
    immediate = not immediate_hits.empty
    if immediate:
        r_rank = "A"

    items = {"腐食": corrosion, "たるみ": sag, "発生率": r_rank}
    n_A = sum(1 for v in items.values() if v == "A")
    n_B = sum(1 for v in items.values() if v == "B")
    any_rank = any(v is not None for v in items.values())

    if n_A >= 2:
        urgency = "Ⅰ"
    elif n_A == 1 or n_B >= 2:
        urgency = "Ⅱ"
    elif any_rank:
        urgency = "Ⅲ"
    else:
        urgency = "異常なし"

    reason = []
    if immediate:
        hits = "、".join(f"{r['item']}a(管No.{int(r['pipe_no'])}, {r['distance_m']}m)"
                         for _, r in immediate_hits.iterrows())
        reason.append(f"{hits} → 特例により発生率ランクA(備考②: 道路陥没等の社会的影響)")
    elif r_rank is not None:
        reason.append(
            f"不良発生率 a={rates['a_pct']}% / a+b={rates['ab_pct']}% / "
            f"a+b+c={rates['abc_pct']}% → 発生率ランク{r_rank}"
        )
    reason.append(f"腐食={corrosion or 'なし'} / たるみ={sag or 'なし'}")
    counts = f"ランクA×{n_A}・B×{n_B}"
    rule = criteria["urgency"].get(urgency, {}).get("rule", "3項目とも異常なし")
    reason.append(f"{counts} → 緊急度{urgency}({rule})")

    return {
        "span_id": span["span_id"],
        "material": span["material"],
        "diameter_mm": span["diameter_mm"],
        "n_pipes": span["n_pipes"],
        "腐食": corrosion or "",
        "たるみ": sag or "",
        "発生率ランク": r_rank or "",
        "a率%": rates["a_pct"], "ab率%": rates["ab_pct"], "abc率%": rates["abc_pct"],
        "特例適用": "○" if immediate else "",
        "緊急度": urgency,
        "判定根拠": " / ".join(reason),
    }


def display_width(text: str) -> int:
    """Width in Excel character units: CJK (and ambiguous) glyphs count double."""
    return sum(2 if unicodedata.east_asian_width(ch) in "FWA" else 1 for ch in text)


def style_sheet(ws, max_width: int = 60) -> None:
    """Make a sheet presentable the way a deliverable workbook is expected to be.

    Header row: filled, white bold, centred. Column width: fitted to the
    widest value counting CJK glyphs as two units (len() would under-size
    Japanese headers). Freeze the header. Print: A4 landscape, all columns
    on one page width, header row repeated on every printed page.
    """
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for idx, col in enumerate(ws.columns, start=1):
        width = max(display_width(str(c.value)) for c in col if c.value is not None)
        ws.column_dimensions[get_column_letter(idx)].width = min(width + 2, max_width)
    ws.freeze_panes = "A2"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = "1:1"


def criteria_sheet_rows(criteria: dict) -> pd.DataFrame:
    """Flatten the criteria profile into a source-annotated table."""
    rows = []
    for item, spec in criteria["span_items"].items():
        if "ranks" in spec:
            for rank, desc in spec["ranks"].items():
                rows.append({"区分": "スパン評価", "項目": item, "管種": "-", "ランク": rank,
                             "基準": desc, "出典": spec["source"]})
        else:
            for band in spec["ranks_by_diameter"]:
                for rank in SPAN_SEVERITY:
                    rows.append({"区分": "スパン評価", "項目": f"{item}(内径{band['diameter_mm_max']}mm以下区分)",
                                 "管種": "-", "ランク": rank, "基準": band[rank], "出典": spec["source"]})
    for item, spec in criteria["pipe_items"].items():
        for mat_key, ranks in spec["ranks"].items():
            mat_name = criteria["materials"][mat_key]
            for rank, desc in ranks.items():
                if desc is None:
                    continue
                rows.append({"区分": "管1本評価", "項目": item, "管種": mat_name, "ランク": rank,
                             "基準": desc, "出典": spec["source"]})
    for grade, spec in criteria["urgency"].items():
        if grade == "source":
            continue
        rows.append({"区分": "緊急度", "項目": f"緊急度{grade}", "管種": "-", "ランク": "-",
                     "基準": f"{spec['rule']}({spec['meaning']})", "出典": criteria["urgency"]["source"]})
    return pd.DataFrame(rows)


def write_report(path: Path, results: pd.DataFrame, defects: pd.DataFrame,
                 errors: pd.DataFrame, criteria: dict) -> None:
    order = ["Ⅰ", "Ⅱ", "Ⅲ", "異常なし"]
    counts = results["緊急度"].value_counts()
    summary = pd.DataFrame(
        [{"項目": "判定実行日", "値": dt.date.today().isoformat()},
         {"項目": "基準プロファイル", "値": criteria["profile_name"]},
         {"項目": "判定スパン数", "値": len(results)},
         *[{"項目": f"緊急度{g}" if g != "異常なし" else g, "値": int(counts.get(g, 0))} for g in order],
         {"項目": "入力エラー件数", "値": len(errors)},
         {"項目": "注意", "値": "本判定は国交省手引き(案)の判定例に準拠したデモであり、自治体の独自基準・維持管理指針本体とは異なる場合があります"}]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="サマリ", index=False)
        results.to_excel(writer, sheet_name="スパン別判定", index=False)
        defects.to_excel(writer, sheet_name="不良明細", index=False)
        criteria_sheet_rows(criteria).to_excel(writer, sheet_name="判定基準と出典", index=False)
        if not errors.empty:
            errors.to_excel(writer, sheet_name="入力エラー", index=False)
        for ws in writer.book.worksheets:
            style_sheet(ws)
        ws = writer.book["スパン別判定"]
        urgency_col = list(results.columns).index("緊急度") + 1
        for row in range(2, len(results) + 2):
            cell = ws.cell(row=row, column=urgency_col)
            color = URGENCY_COLORS.get(str(cell.value))
            if color:
                cell.fill = PatternFill("solid", fgColor=color)
                cell.font = Font(color="FFFFFF", bold=True)


def run(span_path: Path, defect_path: Path, criteria_path: Path,
        out_csv: Path, out_xlsx: Path) -> pd.DataFrame:
    criteria = load_criteria(criteria_path)
    span_df = pd.read_csv(span_path)
    defect_df = pd.read_csv(defect_path)

    errors = validate(span_df, defect_df, criteria)
    valid = defect_df.drop(index=errors["row"].tolist(), errors="ignore")

    results = pd.DataFrame([
        judge_span(span, valid[valid["span_id"] == span["span_id"]], criteria)
        for _, span in span_df.iterrows()
    ])

    out_csv.parent.mkdir(exist_ok=True)
    results.to_csv(out_csv, index=False, encoding="utf-8-sig")
    write_report(out_xlsx, results, defect_df, errors, criteria)
    return results


def main() -> None:
    parser = argparse.ArgumentParser(description="Sewer span urgency judgment (H21 guideline example)")
    parser.add_argument("--criteria", default=str(HERE / "criteria" / "default_h21.json"))
    parser.add_argument("--span", default=str(HERE / "span_master.csv"))
    parser.add_argument("--defect", default=str(HERE / "defect_records.csv"))
    parser.add_argument("--out-csv", default=str(HERE / "output" / "span_urgency.csv"))
    parser.add_argument("--out-xlsx", default=str(HERE / "output" / "urgency_report.xlsx"))
    args = parser.parse_args()

    results = run(Path(args.span), Path(args.defect), Path(args.criteria),
                  Path(args.out_csv), Path(args.out_xlsx))
    print(results["緊急度"].value_counts().reindex(["Ⅰ", "Ⅱ", "Ⅲ", "異常なし"], fill_value=0).to_string())
    print(f"csv  -> {args.out_csv}")
    print(f"xlsx -> {args.out_xlsx}")


if __name__ == "__main__":
    main()
