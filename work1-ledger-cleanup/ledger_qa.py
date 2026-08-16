"""Automated QA inspection for pipeline-ledger GIS layers.

Scans the Handa open-data layers (SonicWeb export) and produces a
data-quality report (Excel + console) covering the checks a ledger
delivery inspection would run:

- file level:   missing/empty .prj (CRS), attribute-definition XML presence
- geometry:     empty geometries, zero-length lines, duplicate SAUID
- attributes:   missing rates of key fields, out-of-range values
                (cover depth, diameter, year), style-ID mixing

Output: qa_report.xlsx (one sheet per layer + summary), console summary.
"""

from __future__ import annotations

import datetime as dt
from dataclasses import dataclass, field
from pathlib import Path

import geopandas as gpd
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parents[1] / "data" / "handa"
OUT_XLSX = Path(__file__).parent / "qa_report.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIS_YEAR = dt.date.today().year  # 暦年基準で判定(年度ではない)

# --- inspection thresholds (with rationale) ---
# 0.01m: below survey precision for pipe segments -> digitizing artifact
MIN_LINE_LENGTH = 0.01
# 5%: missing rate above which attribute completeness needs investigation
MISSING_RATE_WARN = 0.05
# 1900: modern water/sewer works in Japan start late Meiji; earlier = suspect
YEAR_MIN = 1900
# 0-15m: cover depth beyond deep-shield sewers is implausible; negative invalid
COVER_RANGE_M = (0.0, 15.0)


# Material families whose Japanese sewer standard has a known first year.
# A construction year BEFORE the standard existed marks the row as suspect.
# 塩ビ系: JSWAS K-1 制定 1974-06-25 (日本下水道協会 規格制定状況一覧、
# 国総研資料第878号 第2章 p.35 でも確認)
MATERIAL_MIN_YEAR = {
    "塩ビ": (1974, "JSWAS K-1 制定(1974)より前の施工年度"),
}


@dataclass
class LayerSpec:
    rel: str
    label: str
    year_col: str | None = None          # ".." = SAFIELD name
    diameter_col: str | None = None
    cover_cols: list[str] = field(default_factory=list)
    diameter_range: tuple[int, int] = (10, 3500)
    material_col: str | None = None      # enables material-vs-year cross check
    key_col: str | None = None           # business key (uniqueness check)
    key_label: str = "業務キー"
    key_scope: str | None = None         # column whose value scopes the key
                                         # ("SASTYLEID3" = first 3 digits)


LAYERS = [
    LayerSpec("suidou/22001/22001.shp", "上水_管路",
              diameter_col="SAFIELD001", cover_cols=["SAFIELD002"]),
    LayerSpec("suidou/22061/22061.shp", "上水_給水管", diameter_col="SAFIELD002"),
    LayerSpec("suidou/22101/22101.shp", "上水_弁栓",
              year_col="SAFIELD024", diameter_col="SAFIELD007",
              cover_cols=["SAFIELD016"]),
    LayerSpec("suidou/22116/22116.shp", "上水_消火栓",
              year_col="SAFIELD024", diameter_col="SAFIELD007",
              cover_cols=["SAFIELD016"]),
    LayerSpec("gesui_osui/24021/24021.shp", "下水_管渠",
              year_col="SAFIELD000", diameter_col="SAFIELD002",
              cover_cols=["SAFIELD003", "SAFIELD004"], diameter_range=(100, 9000),
              material_col="SAFIELD001"),
    LayerSpec("gesui_osui/24001/24001.shp", "下水_マンホール",
              year_col="SAFIELD018",
              key_col="SAFIELD002", key_label="人孔キー", key_scope="SASTYLEID3"),
    LayerSpec("gesui_osui/24051/24051.shp", "下水_取付管",
              diameter_col="SAFIELD001", diameter_range=(50, 300)),
    LayerSpec("gesui_osui/24041/24041.shp", "下水_桝",
              diameter_col="SAFIELD001", diameter_range=(50, 300)),
]


def check_layer(spec: LayerSpec) -> list[dict]:
    path = BASE / spec.rel
    findings: list[dict] = []

    def add(check: str, result: str, count: int | str, note: str = "") -> None:
        findings.append({"レイヤ": spec.label, "検査項目": check,
                         "判定": result, "件数": count, "備考": note})

    prj = path.with_suffix(".prj")
    if not prj.exists() or prj.stat().st_size == 0:
        add("座標系定義(.prj)", "NG", 1, "空または欠落。座標範囲からの推定と検証が必要")
    else:
        add("座標系定義(.prj)", "OK", 0)

    xml = path.with_suffix(".xml")
    add("属性定義XML", "OK" if xml.exists() else "NG",
        0 if xml.exists() else 1,
        "" if xml.exists() else "SAFIELD列の意味が特定不能")

    gdf = gpd.read_file(path, encoding="cp932")
    n = len(gdf)
    add("レコード数", "-", n)

    empty = gdf.geometry.is_empty | gdf.geometry.isna()
    add("空ジオメトリ", "NG" if empty.any() else "OK", int(empty.sum()))

    lines = gdf[gdf.geom_type.isin(["LineString", "MultiLineString"])]
    if len(lines):
        zero = lines.geometry.length < MIN_LINE_LENGTH
        add("延長ほぼゼロの線", "NG" if zero.any() else "OK", int(zero.sum()),
            f"{MIN_LINE_LENGTH}(座標単位)未満。CRS未定義時は単位に注意")

    if "SAUID" in gdf.columns:
        dup = gdf.duplicated(subset="SAUID", keep=False)
        add("SAUID重複", "NG" if dup.any() else "OK", int(dup.sum()))
    else:
        add("ID列(SAUID)", "NG", 1, "ID列が存在しない")

    if "SASTYLEID" in gdf.columns:
        styles = gdf["SASTYLEID"].nunique()
        add("表示様式(SASTYLEID)の混在", "要確認" if styles > 1 else "OK", styles,
            "複数様式が1ファイルに混在。分析前に分類が必要" if styles > 1 else "")

    if spec.year_col and spec.year_col in gdf.columns:
        years = pd.to_numeric(
            gdf[spec.year_col].astype(str).str.extract(r"(\d{4})")[0],
            errors="coerce")
        miss = years.isna()
        add("年度の欠損", "要確認" if miss.mean() > MISSING_RATE_WARN else "OK",
            int(miss.sum()), f"欠損率 {miss.mean():.1%}")
        bad = years.notna() & ~years.between(YEAR_MIN, THIS_YEAR)
        add("年度の範囲外", "NG" if bad.any() else "OK", int(bad.sum()),
            f"{YEAR_MIN}年未満または未来(暦年基準)")

    if spec.diameter_col and spec.diameter_col in gdf.columns:
        diam = pd.to_numeric(gdf[spec.diameter_col], errors="coerce")
        lo, hi = spec.diameter_range
        miss = diam.isna()
        add("口径の欠損", "要確認" if miss.mean() > MISSING_RATE_WARN else "OK",
            int(miss.sum()), f"欠損率 {miss.mean():.1%}")
        zero = diam == 0
        add("口径の未入力(0)", "要確認" if zero.any() else "OK", int(zero.sum()),
            "0 は未入力プレースホルダとみなす")
        over = diam.notna() & (diam > hi)
        add("口径の上限超過", "要確認" if over.any() else "OK", int(over.sum()),
            f"{hi}mm 超。大型構造物の実在可能性あり、現地資料と照合")
        under = diam.notna() & (diam < lo) & ~zero
        add("口径の下限未満(入力疑義)", "NG" if under.any() else "OK",
            int(under.sum()), f"{lo}mm 未満(0 を除く)。桁誤り等の疑い")

    # --- cross check: material family vs construction year ---------------
    # Born from the series: a 1955 PVC pipe surfaced while writing about the
    # PVC generation shift. A material cannot predate its product standard.
    if spec.material_col and spec.material_col in gdf.columns and spec.year_col:
        years = pd.to_numeric(
            gdf[spec.year_col].astype(str).str.extract(r"(\d{4})")[0],
            errors="coerce")
        mat = gdf[spec.material_col].astype(str)
        for family, (min_year, why) in MATERIAL_MIN_YEAR.items():
            hit = mat.str.contains(family, na=False) & years.notna() & (years < min_year)
            add(f"管種×年代の整合({family}系)", "NG" if hit.any() else "OK",
                int(hit.sum()), why + "。桁誤り・管種誤記・更生の未反映等の疑い")

    # --- business-key uniqueness, scoped and unscoped --------------------
    # Born from the series: manhole keys turned out to be unique only WITHIN
    # a drainage system (2,505 cross-system pairs, 2 real duplicates). A raw
    # duplicate count hides that structure, so report both.
    if spec.key_col and spec.key_col in gdf.columns:
        key = gdf[spec.key_col].astype("string")
        n_blank = int(key.isna().sum())
        add(f"{spec.key_label}の欠損", "NG" if n_blank else "OK", n_blank,
            "施設を特定するキーが空")
        nn = key.dropna()
        n_dup = int(nn.duplicated().sum())
        note = ""
        if spec.key_scope == "SASTYLEID3" and "SASTYLEID" in gdf.columns:
            scope = gdf.loc[nn.index, "SASTYLEID"].astype(str).str[:3]
            scoped = scope + "|" + nn
            n_dup_scoped = int(scoped.duplicated().sum())
            note = (f"全体重複 {n_dup} のうち系統内の真の重複は {n_dup_scoped}。"
                    "残りは系統別採番(キーは系統内でのみ一意)によるもの")
            add(f"{spec.key_label}の重複(系統内)",
                "NG" if n_dup_scoped else "OK", n_dup_scoped, note)
        add(f"{spec.key_label}の重複(全体)",
            "要確認" if n_dup else "OK", n_dup,
            "系統をまたぐ再利用を含む" if note else "")

    for col in spec.cover_cols:
        if col not in gdf.columns:
            continue
        cover = pd.to_numeric(gdf[col], errors="coerce")
        miss = cover.isna()
        add(f"土被りの欠損({col})",
            "要確認" if miss.mean() > MISSING_RATE_WARN else "OK",
            int(miss.sum()), f"欠損率 {miss.mean():.1%}")
        lo_c, hi_c = COVER_RANGE_M
        bad = cover.notna() & ~cover.between(lo_c, hi_c)
        add(f"土被り異常値({col})", "NG" if bad.any() else "OK", int(bad.sum()),
            f"許容 {lo_c}-{hi_c}m。負値・桁違いを検出")

    return findings


def style_sheet(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for idx, col in enumerate(ws.columns, start=1):
        width = max(len(str(c.value)) for c in col if c.value is not None)
        ws.column_dimensions[get_column_letter(idx)].width = min(width + 4, 50)
    ws.freeze_panes = "A2"


def main() -> None:
    all_findings: list[dict] = []
    for spec in LAYERS:
        print(f"checking {spec.label} ...")
        all_findings.extend(check_layer(spec))

    report = pd.DataFrame(all_findings)
    ng = report[report["判定"] == "NG"]
    warn = report[report["判定"] == "要確認"]

    summary = pd.DataFrame([
        {"項目": "検査レイヤ数", "値": len(LAYERS)},
        {"項目": "検査項目数", "値": len(report)},
        {"項目": "NG 検出", "値": len(ng)},
        {"項目": "要確認", "値": len(warn)},
        {"項目": "検査基準日", "値": f"{THIS_YEAR}年"},
    ])

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="サマリ", index=False)
        report.to_excel(writer, sheet_name="全検査結果", index=False)
        ng.to_excel(writer, sheet_name="NG一覧", index=False)
        for ws in writer.book.worksheets:
            style_sheet(ws)

    print(f"\nNG {len(ng)} 件 / 要確認 {len(warn)} 件")
    print(ng.to_string(index=False))
    print(f"\nreport -> {OUT_XLSX}")


if __name__ == "__main__":
    main()
