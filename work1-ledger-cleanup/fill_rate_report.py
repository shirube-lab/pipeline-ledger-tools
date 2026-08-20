"""Attribute fill-rate audit across ALL distributed Handa layers.

ledger_qa.py checks whether entered values are plausible; this script asks
the question one step earlier: is anything entered at all?

Three outputs (fill_rate_report.xlsx):
- レイヤ棚卸し : all 34 distributed files — records, geometry, attribute
                 columns, whether an attribute dictionary exists, and
                 whether the file duplicates another folder's file
                 (same data, only the SAUPDATE export timestamp differs)
- 属性記入率   : per-attribute fill rate for the attribute-bearing layers,
                 with Japanese names restored from the XML dictionary
- サマリ       : headline numbers for the article

"Filled" = not null and not blank/whitespace after strip. Zero counts as
filled (placeholder zeros are a value-plausibility problem — ledger_qa's
job — not a fill problem); numeric columns also get a zero-rate column so
placeholder-suspect cases stay visible.
"""

from __future__ import annotations

from pathlib import Path

import geopandas as gpd
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parents[1] / "data" / "handa"
DICT_CSV = Path(__file__).parent / "s2a_field_mapping.csv"
OUT_XLSX = Path(__file__).parent / "fill_rate_report.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)

# columns every SonicWeb export carries; not business attributes
SYSTEM_COLS = {"SAUID", "SAUPDATE", "SADATEST", "SADATEED", "SASTYLEID",
               "SAANGLE", "SASIZE", "SATEXT", "SAVERTICAL", "geometry"}


def filled(series: pd.Series) -> pd.Series:
    """True where a real value is present (not null, not blank text)."""
    if series.dtype == object or pd.api.types.is_string_dtype(series):
        s = series.astype("string")
        return s.notna() & (s.str.strip().str.len() > 0)
    return series.notna()


def main() -> None:
    fdict = pd.read_csv(DICT_CSV)
    titles = {(g, str(no), fid): t for g, no, fid, t in
              fdict[["group", "layer_no", "field_id", "title"]].itertuples(index=False)
              if pd.notna(t)}

    inventory: list[dict] = []
    frames: dict[tuple[str, str], gpd.GeoDataFrame] = {}
    for shp in sorted(BASE.rglob("*.shp")):
        grp, code = shp.parent.parent.name, shp.parent.name
        g = gpd.read_file(shp, encoding="cp932")
        frames[(grp, code)] = g
        sa = [c for c in g.columns if c.startswith("SAFIELD")]
        kind = ("業務属性あり" if sa
                else "注記(文字列)" if filled(g.get("SATEXT", pd.Series(dtype=object))).any()
                else "図形・記号のみ")
        inventory.append({
            "フォルダ": grp, "レイヤ番号": code, "レコード数": len(g),
            "図形": g.geom_type.iloc[0] if len(g) else "-",
            "業務属性列数": len(sa),
            "属性辞書": "あり" if sa else "なし",
            "分類": kind,
            "レイヤ名": {"24001": "マンホール", "24021": "管渠", "24041": "桝",
                     "24051": "取付管", "22001": "管路", "22061": "給水管",
                     "22101": "弁栓", "22116": "消火栓"}.get(code, "(配布物に名称なし)"),
        })

    # folder duplicates: same code in both sewer folders, same data apart
    # from the SAUPDATE export timestamp
    inv = pd.DataFrame(inventory)
    dup_note: dict[tuple[str, str], str] = {}
    osui = {c for f, c in frames if f == "gesui_osui"}
    usui = {c for f, c in frames if f == "gesui_usui"}
    for code in sorted(osui & usui):
        a = frames[("gesui_osui", code)].drop(columns="geometry")
        b = frames[("gesui_usui", code)].drop(columns="geometry")
        same = False
        if list(a.columns) == list(b.columns) and len(a) == len(b):
            drop = [c for c in ("SAUPDATE",) if c in a.columns]
            aa = a.drop(columns=drop).sort_values("SAUID").reset_index(drop=True)
            bb = b.drop(columns=drop).sort_values("SAUID").reset_index(drop=True)
            same = aa.equals(bb)
        note = ("汚水フォルダと同一データ(差はエクスポート日時列のみ)" if same
                else "汚水フォルダと同番号だが内容差あり(要確認)")
        if same:
            dup_note[("gesui_usui", code)] = note
    inv["重複"] = [dup_note.get((r["フォルダ"], r["レイヤ番号"]), "")
                  for _, r in inv.iterrows()]

    # fill rates for unique attribute-bearing layers (skip usui duplicates)
    rate_rows: list[dict] = []
    for (grp, code), g in frames.items():
        sa_cols = [c for c in g.columns if c.startswith("SAFIELD")]
        if not sa_cols or (grp, code) in dup_note:
            continue
        label = inv[(inv["フォルダ"] == grp) & (inv["レイヤ番号"] == code)]["レイヤ名"].iloc[0]
        for c in sa_cols:
            f = filled(g[c])
            num = pd.to_numeric(g[c], errors="coerce")
            zero = int((num == 0).sum()) if num.notna().any() else 0
            rate_rows.append({
                "レイヤ": f"{label}({code})", "列": c,
                "属性名": titles.get((grp, code, c), "(辞書に無し)"),
                "記入率": round(f.mean() * 100, 1),
                "記入数": int(f.sum()), "全数": len(g),
                "うち0の数": zero,
            })
    rates = pd.DataFrame(rate_rows)

    summary = pd.DataFrame([
        {"項目": "配布ファイル数", "値": len(inv)},
        {"項目": "実質レイヤ数(重複を除く)", "値": len(inv) - len(dup_note)},
        {"項目": "うち業務属性を持つレイヤ", "値": int((inv["重複"] == "").mul(inv["業務属性列数"] > 0).sum())},
        {"項目": "検査した属性数", "値": len(rates)},
        {"項目": "記入率90%以上の属性", "値": int((rates["記入率"] >= 90).sum())},
        {"項目": "記入率10%未満の属性", "値": int((rates["記入率"] < 10).sum())},
        {"項目": "記入率0%(全件空)の属性", "値": int((rates["記入率"] == 0).sum())},
    ])

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="サマリ", index=False)
        inv.to_excel(writer, sheet_name="レイヤ棚卸し", index=False)
        rates.to_excel(writer, sheet_name="属性記入率", index=False)
        for ws in writer.book.worksheets:
            for cell in ws[1]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")
            for idx, col in enumerate(ws.columns, start=1):
                width = max(len(str(c.value)) for c in col if c.value is not None)
                ws.column_dimensions[get_column_letter(idx)].width = min(width + 4, 60)
            ws.freeze_panes = "A2"

    print(summary.to_string(index=False))
    print("\n--- 記入率ワースト(0%を除く下位10) ---")
    nz = rates[rates["記入率"] > 0].nsmallest(10, "記入率")
    print(nz.to_string(index=False))
    print(f"\n0%(全件空)の属性: {int((rates['記入率'] == 0).sum())} 本")
    print(rates[rates["記入率"] == 0][["レイヤ", "列", "属性名"]].to_string(index=False))
    print(f"\nreport -> {OUT_XLSX}")


if __name__ == "__main__":
    main()
