"""Reconcile a GIS pipe ledger against field survey results.

Reads the ledger (GeoJSON, via geopandas) and the survey (CSV), runs a
set of reconciliation and validation checks, and writes an Excel report
(check_report.xlsx) with one sheet per check category.

Usage:
    python ledger_checker.py [--ledger ledger.geojson] [--survey survey.csv]
                             [--out check_report.xlsx]

Note: the future-year check uses the calendar year, not the Japanese
fiscal year.
"""

from __future__ import annotations

import argparse
import datetime as dt
import unicodedata
from pathlib import Path

import geopandas as gpd
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

HERE = Path(__file__).parent
VALID_MATERIALS = {"DIP", "VP", "CIP", "SP", "PE"}
VALID_DIAMETERS = {50, 75, 100, 150, 200, 250, 300, 400}
NUMERIC_COMPARE = ["install_year", "diameter_mm"]
STRING_COMPARE = ["material"]
REQUIRED_COLS = ["install_year", "material", "diameter_mm", "length_m"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def load_inputs(ledger_path: str, survey_path: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    ledger = gpd.read_file(ledger_path)
    ledger = pd.DataFrame(ledger.drop(columns="geometry"))
    survey = pd.read_csv(survey_path)
    return ledger, survey


def check_id_diff(ledger: pd.DataFrame, survey: pd.DataFrame) -> pd.DataFrame:
    """Pipes present on only one side of the reconciliation."""
    l_ids = set(ledger["pipe_id"])
    s_ids = set(survey["pipe_id"])
    rows = [
        {"pipe_id": pid, "issue": "台帳のみ(調査で未確認)"} for pid in sorted(l_ids - s_ids)
    ] + [
        {"pipe_id": pid, "issue": "調査のみ(台帳に未登録)"} for pid in sorted(s_ids - l_ids)
    ]
    return pd.DataFrame(rows, columns=["pipe_id", "issue"])


def check_attr_mismatch(ledger: pd.DataFrame, survey: pd.DataFrame) -> pd.DataFrame:
    """Attribute differences for pipes present on both sides.

    Numeric columns are compared as numbers (so int32 vs int64 or a
    float-promoted "1988.0" never causes false mismatches); string
    columns are compared after stripping whitespace. NaN on both sides
    counts as a match — missing values are reported by check_validity.
    """
    merged = ledger.merge(survey, on="pipe_id", suffixes=("_ledger", "_survey"))
    rows = []

    def report(diff: pd.DataFrame, col: str) -> None:
        a, b = f"{col}_ledger", f"{col}_survey"
        for _, r in diff.iterrows():
            rows.append(
                {"pipe_id": r["pipe_id"], "attribute": col,
                 "ledger_value": r[a], "survey_value": r[b]}
            )

    for col in NUMERIC_COMPARE:
        a = pd.to_numeric(merged[f"{col}_ledger"], errors="coerce")
        b = pd.to_numeric(merged[f"{col}_survey"], errors="coerce")
        mismatch = (a != b) & ~(a.isna() & b.isna())
        report(merged[mismatch], col)

    for col in STRING_COMPARE:
        a = merged[f"{col}_ledger"].astype("string").str.strip()
        b = merged[f"{col}_survey"].astype("string").str.strip()
        mismatch = (a != b) & ~(a.isna() & b.isna())
        report(merged[mismatch], col)

    return pd.DataFrame(rows, columns=["pipe_id", "attribute", "ledger_value", "survey_value"])


def check_validity(ledger: pd.DataFrame) -> pd.DataFrame:
    """Standalone validation of ledger values."""
    this_year = dt.date.today().year
    rows = []

    def add(df: pd.DataFrame, issue: str, value_col: str) -> None:
        for _, r in df.iterrows():
            rows.append({"pipe_id": r["pipe_id"], "issue": issue, "value": r[value_col]})

    # Missing required values first; range checks below only see filled rows
    for col in REQUIRED_COLS:
        add(ledger[ledger[col].isna()], f"必須項目の未入力({col})", col)
    filled = ledger.dropna(subset=REQUIRED_COLS)

    add(filled[filled["install_year"] > this_year], "布設年度が未来", "install_year")
    add(filled[filled["install_year"] < 1900], "布設年度が異常に古い", "install_year")
    add(filled[~filled["diameter_mm"].isin(VALID_DIAMETERS)], "口径が規格外", "diameter_mm")
    add(filled[filled["length_m"] <= 0], "延長が0以下", "length_m")
    add(filled[~filled["material"].isin(VALID_MATERIALS)], "管種の表記が不正", "material")

    dup = ledger[ledger.duplicated(subset="pipe_id", keep=False)]
    add(dup, "pipe_id が重複", "pipe_id")

    return pd.DataFrame(rows, columns=["pipe_id", "issue", "value"])


def display_width(text: str) -> int:
    """Width in Excel character units: CJK (and ambiguous) glyphs count double."""
    return sum(2 if unicodedata.east_asian_width(ch) in "FWA" else 1 for ch in text)


def style_sheet(ws, max_width: int = 40) -> None:
    """Make a sheet presentable the way a deliverable workbook is expected to be.

    Header row: filled, white bold, centred. Column width: fitted to the
    widest value counting CJK glyphs as two units (len() would under-size
    Japanese headers). Freeze the header. Print: A4 landscape, all columns
    on one page width, header row repeated on every printed page.
    """
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for idx, col in enumerate(ws.columns, start=1):
        width = max(display_width(str(c.value)) for c in col if c.value is not None)
        ws.column_dimensions[get_column_letter(idx)].width = min(width + 2, max_width)
        if width + 2 > max_width:            # long text: wrap instead of clipping
            for c in col[1:]:
                c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = "1:1"
    ws.print_options.gridLines = True   # screen gridlines are NOT printed by default


def write_report(path: str, ledger: pd.DataFrame, survey: pd.DataFrame,
                 id_diff: pd.DataFrame, mismatch: pd.DataFrame,
                 validity: pd.DataFrame) -> None:
    summary = pd.DataFrame(
        [
            {"項目": "チェック実行日", "値": dt.date.today()},
            {"項目": "台帳レコード数", "値": len(ledger)},
            {"項目": "調査レコード数", "値": len(survey)},
            {"項目": "ID差分件数", "値": len(id_diff)},
            {"項目": "属性不一致件数", "値": len(mismatch)},
            {"項目": "妥当性エラー件数", "値": len(validity)},
        ]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="サマリ", index=False)
        id_diff.to_excel(writer, sheet_name="ID差分", index=False)
        mismatch.to_excel(writer, sheet_name="属性不一致", index=False)
        validity.to_excel(writer, sheet_name="妥当性エラー", index=False)
        for ws in writer.book.worksheets:
            style_sheet(ws)


def main() -> None:
    parser = argparse.ArgumentParser(description="Pipe ledger reconciliation checker")
    parser.add_argument("--ledger", default=str(HERE / "ledger.geojson"))
    parser.add_argument("--survey", default=str(HERE / "survey.csv"))
    parser.add_argument("--out", default=str(HERE / "check_report.xlsx"))
    args = parser.parse_args()

    ledger, survey = load_inputs(args.ledger, args.survey)
    id_diff = check_id_diff(ledger, survey)
    mismatch = check_attr_mismatch(ledger, survey)
    validity = check_validity(ledger)
    write_report(args.out, ledger, survey, id_diff, mismatch, validity)

    print(f"ID差分: {len(id_diff)} 件 / 属性不一致: {len(mismatch)} 件 / "
          f"妥当性エラー: {len(validity)} 件")
    print(f"report -> {args.out}")


if __name__ == "__main__":
    main()
